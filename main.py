from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from tabulate import tabulate
import os

wb = load_workbook('bank.xlsx')
ws = wb.active

def does_value_exist(value,col1,col2):
  count = 0
  uid = value
  end_range = len(ws['A'])+1
  for row in range(2,end_range):
    for col in range(col1,col2):
      char = get_column_letter(col)
      if (ws[char + str(row)].value) == uid:
        cell_row = row
        count +=1
        break
    if (ws[char + str(row)].value) == uid:
      break
  if count == 1:
    count = 0
    return cell_row
  else:
    return False

def pin_validity(pin,cell_row):
  for row in range(cell_row,cell_row+1):
    for col in range(6,7):
      char = get_column_letter(col)
      if str(ws[char + str(row)].value) == (pin):
        return "valid"
      else:
        return "invalid"  

def account_details(cell_row):
  title_list=[]
  details_list =[]

  for row in range(1,2):
    for col in range(2,6):
      char = get_column_letter(col)
      title_list.append((ws[char + str(row)].value))

  for row in range(cell_row,cell_row+1):
    for col in range(2,6):
      char = get_column_letter(col)
      details_list.append((ws[char + str(row)].value))

  combined_list = [list(l) for l in zip(title_list, details_list)]
  
  print(tabulate(combined_list))
  return details_list[2]

def commandList(account_type):
  if account_type == "general user":
    print("-------------------------\n"
      "What do you want to do :\n"
      "(1) Deposit balance\n"
      "(2) Withdraw balance\n"
      "(3) Transfer balance\n"
      "(4) See Status\n"
      "(5) Exit\n"
      "-------------------------\n"
    )
  elif account_type == "admin":
    print(
      "-------------------------\n"
      "What do you want to do :\n"
      "(1) Add Data\n"
      "(2) Remove Data\n"
      "(3) Edit Data\n"
      "(4) Exit\n"
      "-------------------------\n")

def get_cell_address(title,row):
  title_values ={
    "Account Number" : 'B' ,
    "User Name"      : 'C' ,  
    "Pin"            : 'F'
  }
  col = title_values[title] 
  return col + str(row)

def Deposit(cell_row,type,transferred):
  if type == "deposit":
    os.system("cls")
    deposit_amt = int(input("How much would you like to deposit?: "))
  elif type == "transfer":
    deposit_amt = transferred
  for row in range(cell_row,cell_row+1):
    for col in range(5,6):
      char = get_column_letter(col)
      previous_bal = ws[char + str(row)].value
      ws[char + str(row)].value = int(previous_bal)+deposit_amt
  new_bal = previous_bal + deposit_amt

  if type == "deposit":    
    print(f"Successfully deposited amount {deposit_amt}\n"
    f"Your new balance is {new_bal}\n")
  wb.save('bank.xlsx')
  
def Withdraw(cell_row,type):
  for row in range(cell_row,cell_row+1):
    for col in range(5,6):
      char = get_column_letter(col)
      previous_bal = ws[char + str(row)].value
      withdraw_amt = int(input(f"How much would you like to {type}?\n"
      f"maximum amount |{previous_bal}|: "))
      if withdraw_amt<= previous_bal:
        ws[char + str(row)].value = int(previous_bal)-withdraw_amt
        new_bal = previous_bal - withdraw_amt
        pronoun = "withdrew" if type == "withdraw" else "transferred"
        print(
          f"Successfully {pronoun} amount {withdraw_amt}\n"
          f"Your new balance is {new_bal}\n")
      else:
        print(f"You dont have enough balance for the {type}")

  wb.save('bank.xlsx')
  return withdraw_amt

def Transfer(cell_row):
  receiver = int(input("Enter account number of the receiver: "))
  receiver_acc_no = does_value_exist(receiver,2,3)
  if receiver_acc_no != False:
    transfer = Withdraw(cell_row,"transfer")
    Deposit(receiver_acc_no,"transfer",transfer)
    
def add_data():
  end_range = len(ws['A'])
  uid = end_range-1
  acc_no = int(input("Enter Account Number: "))
  cell_row = does_value_exist(acc_no,2,3)
  if cell_row != False:
    print("The account number exists please create unique account number\n")
    commandList("admin")
  else:
    user_name = input("Enter user name: ")
    account_type = input("Enter account type: ")
    balance = int(input("Enter balance: "))
    pin = input("Enter pin value: ")
    ws.append([uid, acc_no, user_name, account_type, balance,pin])
    wb.save('bank.xlsx')
    print("Data was stored in the database successfully\n")

def remove_data():
  uid_to_remove = int(input("Enter the UID whose data is to be removed: "))
  does_uid_exists = does_value_exist(uid_to_remove,1,2)
  if does_uid_exists == False:
    print("The UID doesnt exists\n")
  else:
    is_admin = account_details(does_uid_exists)
    if is_admin == "admin":
      os.system("cls")
      print("Only owner can delete admin accounts\n")   
    else:
      is_delete = input(
        "! Warning the above data will be erased..\n"
        "do you wish to proceeed Y/N: ").lower()
      if is_delete == 'y':
        os.system("cls")
        ws.delete_rows(does_uid_exists)
        print("Successfully deleted the data from the database\n")
        wb.save('bank.xlsx')
      else:
        os.system("cls")

def edit_data():
  uid = int(input("Enter UID for edit: "))
  does_uid_exists=(does_value_exist(uid,1,2))
  if does_uid_exists == False:
    print("The UID doesnt exists\n")
  else:
    print("...previous value...\n")
    account_details(does_uid_exists)
    acc_no = int(input("Enter Account Number: "))
    does_acc_no_exists = does_value_exist(acc_no,2,3)
    if does_acc_no_exists != False:
      print("The account number exists please create unique account number\n")
      commandList("admin")
    else:
      acc_no_address = get_cell_address("Account Number",does_uid_exists)
      ws[acc_no_address] = acc_no
      user_name = input("Enter user name: ")
      user_name_address = get_cell_address("User Name",does_uid_exists)
      ws[user_name_address] = user_name
      pin = input("Enter pin value: ")
      pin_address = get_cell_address("Pin",does_uid_exists)
      ws[pin_address] = pin
      os.system("cls")
      print("Data Successfully Edited\n"
      "...New data is...\n")
      wb.save('bank.xlsx')
      account_details(does_uid_exists)
      
def available_function(account_type,cell_row):
  if account_type == "general user":
    commandList("general user")
    while True:
      command = input("command number: ")
      if command == "1":
        Deposit(cell_row,"deposit",0)
        commandList("general user")
      elif command == "2":
        os.system("cls")
        Withdraw(cell_row,"withdraw")
        commandList("general user")
      elif command == "3":
        os.system("cls")
        Transfer(cell_row)
        commandList("general user")
      elif command == "4":
        os.system("cls")
        print("User details\n")
        account_details(cell_row)
        commandList("general user")
      else:  
        break

  else:
    commandList("admin")
    while True:
      command = input("command number: ")
      if command == "1":
        os.system("cls")
        add_data()
        commandList("admin")
      elif command == "2":
        os.system("cls")
        remove_data()
        commandList("admin")
      elif command == "3":
        os.system("cls")
        edit_data()
        commandList("admin")
      else:  
        break
    
print(
  "---------------\n"
  "Welcome to bank\n"
  "---------------\n"
)

def main():
  uid = int(input("Enter you ID: "))
  cell_row = (does_value_exist(uid,1,2))
  if cell_row != False:
    pin =input("Enter pin: ")
    if (pin_validity(pin,cell_row)) == "valid":
      print("\nUser details")
      account_type = account_details(cell_row)
      available_function(account_type,cell_row)
      is_continue = input("Do you want to continue? Y/N: ").lower()
      if(is_continue == 'y'):
        os.system("cls")
        main()
      else:
        exit()
    else:
      print("Pin doesnt match\n")
      main()

  else:
    print("UID doesnt exists in the databse\n"
      "contact admins or try another UID\n")
    is_exit = int(input("Enter 0 to exit: "))
    if is_exit == 0:
      exit()
    else:
      os.system("cls")
      main()

main()
